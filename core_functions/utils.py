from typing import Optional
import os, hashlib
import pandas as pd
import numpy as np
import pickle, json
from core_functions import config
from openpyxl.styles import Alignment, Font, PatternFill

### КЛАССЫ ###
class DirectoriesManager:
    def __init__(self, structure: dict, root_path: str = ".", build: bool=True):
        self.paths = {}
        self._build_structure(structure, root_path, build)

    def _build_structure(self, current_structure: dict, current_path: str, build: bool):
        for folder_name, subfolders in current_structure.items():
            folder_path = os.path.join(current_path, folder_name)
            if os.path.isdir(folder_path):
                if build: print_info(f'Найдена папка: {folder_path}')
            else:
                os.makedirs(folder_path, exist_ok=False)
                if build: print_info(f"Создана папка: {folder_path}")
            self.paths[folder_name] = folder_path

            if isinstance(subfolders, dict):
                self._build_structure(subfolders, folder_path, build)

    def get_dir_path(self, folder_name: str) -> str:
        return self.paths.get(folder_name, "")
    
class CSVManager:
    def __init__(self, file_name: str, file_column: str):
        self.file_name = file_name
        self.file_column = file_column

    """ ПРИВАТНЫЕ МЕТОДЫ """
    def _check_file_and_column(self, file_path: str, file_extension: str) -> bool:
        if file_extension not in {'.xlsb', '.csv'}:
            raise ValueError("file_extension может принимать только значения '.xlsb' и '.csv'.\n")
        
        if not file_path.lower().endswith(file_extension):
            print_critical_error(f"Расширение файла должно быть: '{file_extension}'", prefix='\n')
            print_critical_error(f"Файл: {self.file_name}", end='\n\n')
            return False

        if not os.path.isfile(file_path):
            if file_extension == '.csv': return True # Сигнал для создания .csv файла
            print_critical_error("Исходный файл с обращениями не найден.", prefix='\n')
            print_critical_error(f"Файл: {file_path}", end='\n\n')
            return False
        
        try:
            if file_extension == '.xlsb':
                df = pd.read_excel(file_path, engine='pyxlsb', nrows=1)
            else: # '.csv'
                df = pd.read_csv(file_path, encoding='utf-8', nrows=1)
            if self.file_column not in df.columns:
                print_critical_error(f"Колонка '{self.file_column}' не найдена в файле.", prefix='\n')
                print_critical_error(f"Файл: {file_path}", end='\n\n')
                return False
        except Exception as e:
            print_critical_error("Невозможно прочитать колонку в файле.", prefix='\n')
            print_critical_error(f"Файл: {file_path}")
            print_critical_error(f"Колонка: {self.file_column}")
            print_critical_error(f"Причина: {e}", end='\n\n')
            return False
        return True
    
    def _get_file_hash(self) -> str:
        file_hash = hashlib.md5(f'{self.file_name}_{self.file_column}'.encode('utf-8')).hexdigest()
        return file_hash

    def _load_data_from_excel(self, dir_path: str) -> list[str]:
        if not os.path.isdir(dir_path):
            print_critical_error('Папка не найдена.', prefix='\n')
            print_critical_error(f'Путь: {dir_path}', end='\n\n')
            exit(1)
            
        file_path = os.path.join(dir_path, self.file_name)
        
        try:
            not_error = self._check_file_and_column(file_path, '.xlsb')
            if not not_error:
                exit(1)
        except Exception as e:
            print_critical_error(f"Невозможно проверить файл '{self.file_name}' и колонку '{self.file_column}'.", prefix='\n')
            print_critical_error(f"Причина: {e}", end='\n\n')
            exit(1)
            
        try:
            print_info('Осуществляется чтение обращений потребителей из исходного .xlsb файла. Пожалуйста, подождите...')
            print_info('Путь к файлу:', end=' '); print_key_info(f'{file_path}')
            df = pd.read_excel(file_path, engine='pyxlsb')
            print_info('Данные обращений потребителей были успешно прочитаны!', end='\n\n')
            return df[self.file_column].dropna().tolist()
        except Exception as e:
            print_critical_error(f'Невозможно прочитать файл.', prefix='\n')
            print_critical_error(f'Путь к файлу: {file_path}')
            print_critical_error(f'Причина: {e}', end='\n\n')
            exit(1)
            
    """ ПУБЛИЧНЫЕ МЕТОДЫ """
    def read_source_data_from_csv(self, dir_path: str) -> Optional[list[str]]:
        if not os.path.isdir(dir_path):
            print_critical_error('Папка не найдена.', prefix='\n')
            print_critical_error(f'Путь: {dir_path}', end='\n\n')
            exit(1)

        subfolder_path = os.path.join(dir_path, os.path.splitext(os.path.basename(self.file_name))[0])
        os.makedirs(subfolder_path, exist_ok=True)
        file_path = os.path.join(subfolder_path, f'{self._get_file_hash()}.csv')

        try:
            not_error = self._check_file_and_column(file_path, '.csv')
            if not not_error:
                exit(1)
        except Exception as e:
            print_critical_error(f"Невозможно проверить файл '{self.file_name}' и колонку '{self.file_column}'.", prefix='\n')
            print_critical_error(f"Причина: {e}", end='\n\n')
            exit(1)

        if os.path.isfile(file_path):
            print_info('Осуществляется чтение обращений потребителей из сохраненного .csv файла. Пожалуйста подождите...')
            print_info('Путь к файлу:', end=' '); print_key_info(f'{file_path}')
            try:    
                df = pd.read_csv(file_path, encoding='utf-8', dtype={self.file_column: str})
                print_info('Данные обращений потребителей были успешно прочитаны!')
                return df[self.file_column].tolist()
            except Exception as e:
                print_critical_error('Невозможно прочитать файл', prefix='\n')
                print_critical_error(f'Путь к файлу: {file_path}')
                print_critical_error(f'Причина: {e}', end='\n\n')
                exit(1)
        else:
            print_warn('Файл .csv с обращениями потребителей не найден.')
            print_warn(f'Путь к файлу: {file_path}', end='\n\n')
            return None
        
    def load_and_save_source_data_to_csv(self, load_dir_path, save_dir_path: str) -> list[str]:
        if not os.path.isdir(load_dir_path):
            print_critical_error('Папка не найдена.', prefix='\n')
            print_critical_error(f'Путь: {load_dir_path}', end='\n\n')
            exit(1)
        
        requests = self._load_data_from_excel(load_dir_path)
        
        if not os.path.isdir(save_dir_path):
            print_critical_error('Папка не найдена.', prefix='\n')
            print_critical_error(f'Путь: {save_dir_path}', end='\n\n')
            exit(1)

        subfolder_path = os.path.join(save_dir_path, os.path.splitext(os.path.basename(self.file_name))[0])    
        os.makedirs(subfolder_path, exist_ok=True)
        file_path = os.path.join(subfolder_path, f'{self._get_file_hash()}.csv')
        
        if not os.path.isfile(file_path):
            print_info('Осуществляется сохранение обращений потребителей в .csv файл. Пожалуйста, подождите...')
            print_info('Путь к файлу:', end=' '); print_key_info(f'{file_path}')
            try:
                df = pd.DataFrame(requests, columns=[self.file_column])
                df.to_csv(file_path, index=False, encoding='utf-8')
                print_info(f'Данные обращений потребителей были успешно сохранены!')
                return requests
            except Exception as e:
                print_critical_error(f'Невозможно сохранить файл.', prefix='\n')
                print_critical_error(f'Путь к файлу: {file_path}')
                print_critical_error(f'Причина: {e}', end='\n\n')
                exit(1)
        else:
            print_warn('Файл .csv с обращениями потребителей уже существует')
            print_warn(f'Путь к файлу: {self.source_data_csv_file_path}')
            print_info(f'Исходный файл: {os.path.join(self.source_data_dir_path, self.file_name)} -> {self.file_column}')
            print_warn('Если вы хотите обработать новый файл, измените его название или колонку.')

    def read_processed_data_from_csv(self, dir_path: str, suffix: str = None) -> tuple[str, Optional[list[str]]]:
        if not os.path.isdir(dir_path):
            print_critical_error(f'Папка не найдена.', prefix='\n')
            print_critical_error(f'Путь: {dir_path}', end='\n\n')
            exit(1)
        
        if suffix is None:
            suffix = '_' + os.path.splitext(self.file_name)[0] + '_' + self.file_column
            
        subfolder_path = os.path.join(dir_path, os.path.splitext(os.path.basename(self.file_name))[0])
        os.makedirs(subfolder_path, exist_ok=True)
        file_path = os.path.join(subfolder_path, f'{self._get_file_hash()}{suffix}.csv')
        
        if os.path.isfile(file_path):
            print_info('Осуществляется чтение предобработанных обращений потребителей из .csv файла. Пожалуйста, подождите...')
            print_info('Путь к файлу:', end=' '); print_key_info(f'{file_path}')
            try:
                df = pd.read_csv(file_path, header=None, encoding='utf-8', low_memory=False).map(lambda x: x if pd.notna(x) else None)
                processed_requests = df.values.tolist()
                print_info('Предобработанные обращения потребителей были успешно прочитаны!')
                return subfolder_path, processed_requests
            except Exception as e:
                print_critical_error('Невозможно прочитать файл.', prefix='\n')
                print_critical_error(f'Путь к файлу: {file_path}')
                print_critical_error(f'Причина: {e}', end='\n\n')
                exit(1)
        else:
            print_warn('Файл .csv с предобработанными обращениями потребителей не найден.')
            print_warn(f'Путь к файлу: {file_path}', end='\n\n')
            return subfolder_path, None
        
    def save_processed_data_to_csv(self, processed_requests: list[list[str]], dir_path: str, suffix: str = None) -> None:
        if not os.path.isdir(dir_path):
            print_critical_error('Папка не найдена.', prefix='\n')
            print_critical_error(f'Путь: {dir_path}', end='\n\n')
            exit(1)
            
        if suffix is None:
            suffix = '_' + os.path.splitext(self.file_name)[0] + '_' + self.file_column

        subfolder_path = os.path.join(dir_path, os.path.splitext(os.path.basename(self.file_name))[0])
        os.makedirs(subfolder_path, exist_ok=True)
        file_path = os.path.join(subfolder_path, f'{self._get_file_hash()}{suffix}.csv')
        
        if not os.path.isfile(file_path):
            print_info('Осуществляется сохранение предобработанных обращений потребителей в .csv файл. Пожалуйста, подождите...')
            print_info('Путь к файлу:', end=' '); print_key_info(f'{file_path}')
            try:
                df = pd.DataFrame(processed_requests)
                df.to_csv(file_path, index=False, header=False, encoding='utf-8')
                print_info(f'Предобработанные обращения потребителей были успешно сохранены!')
            except Exception as e:
                print_critical_error('Невозможно сохранить файл.', prefix='\n')
                print_critical_error(f'Путь к файлу: {file_path}')
                print_critical_error(f'Причина: {e}', end='\n\n')
                exit(1)
        else:
            print_warn('Файл .csv с предобработанными обращениями потребителей уже существует.')
            print_warn(f'Путь к файлу: {self.preprocessed_source_data_csv_file_path}')
            print_info(f'Исходный файл: {os.path.join(self.source_data_dir_path, self.file_name)} -> {self.file_column}')
            print_warn('Если вы хотите обработать новый файл, измените его название или колонку.')
            
    def get_file_hash(self):
        return self._get_file_hash()

### ЦВЕТОВАЯ ГАММА ###
ANSI = {
    'RESET': '\033[0m',
    'WELCOME_SIGN': '\033[38;5;245m',
    'WELCOME_TEXT': '\033[38;5;109m',
    'SUCCESS': '\033[38;5;35m',
    'WARN': '\033[38;5;214m',
    'ERROR': '\033[38;5;196m',
    'CRITICAL_ERROR': '\033[38;5;124m',
    'INFO': '\033[38;5;32m',
    'KEY_INFO' : '\033[38;5;33m',
    'MENU_OPTION': '\033[38;5;60m',
    'MAIN_LINE': '\033[38;5;241m',
    'SUB_LINE': '\033[38;5;245m',
    'INSCRIPTION': '\033[38;5;103m'
    }

### ФУНКЦИИ ВЫВОДА ###
def print_welcome(width: int=108) -> None:
    equals = width - 2
    text = 'ДОБРО ПОЖАЛОВАТЬ В ПРОГРАММУ'
    left = (equals - len(text)) // 2
    right = equals - len(text) - left
    print()
    print(f"{ANSI['WELCOME_SIGN']}|{'=' * equals}|{ANSI['RESET']}")
    print(f"{ANSI['WELCOME_SIGN']}|{ANSI['RESET']}{' '*left}{ANSI['WELCOME_TEXT']}{text}{ANSI['RESET']}{' '*right}{ANSI['WELCOME_SIGN']}|{ANSI['RESET']}")
    print(f"{ANSI['WELCOME_SIGN']}|{'=' * equals}|{ANSI['RESET']}")
    print()
    
def print_success(text: str, end='\n', prefix: str='') -> None:
    print(f"{prefix}{ANSI['SUCCESS']}[SUCCESS]{ANSI['RESET']} {text}", end=end)

def print_warn(text: str, end='\n', prefix: str='') -> None:
    print(f"{prefix}{ANSI['WARN']}[WARN]{ANSI['RESET']} {text}", end=end)

def print_error(text: str, end='\n', prefix: str='') -> None:
    print(f"{prefix}{ANSI['ERROR']}[ERROR]{ANSI['RESET']} {text}", end=end)

def print_critical_error(text: str, end='\n', prefix: str='') -> None:
    print(f"{prefix}{ANSI['CRITICAL_ERROR']}[CRITICAL_ERROR]{ANSI['RESET']} {text}", end=end)

def print_info(text: str, end='\n', prefix: str='') -> None:
    print(f"{prefix}{ANSI['INFO']}[INFO]{ANSI['RESET']} {text}", end=end)
 
def print_key_info(text: str, end='\n', prefix: str='') -> None:
    print(f"{prefix}{ANSI['KEY_INFO']}{text}{ANSI['RESET']}", end=end)
   
def print_menu_option(text: str, end='\n', prefix: str='') -> None:
    print(f"{prefix}{ANSI['MENU_OPTION']}{text}{ANSI['RESET']}", end=end)
    
def print_inscription(text: str, end='\n', width: int=108) -> None:
    total_len = len(text) + 2
    dashes = width - total_len
    left = dashes // 2
    right = dashes - left
    print(f"{ANSI['INSCRIPTION']}{'-'*left}>{text}<{'-'*right}{ANSI['RESET']}", end=end)
    
def print_main_line(width: int=108) -> None:
    print(f"{ANSI['MAIN_LINE']}{'+' * width}{ANSI['RESET']}")
    
def print_sub_line(width: int=108) -> None:
    print(f"{ANSI['SUB_LINE']}{'~' * width}{ANSI['RESET']}")

### ФУНКЦИИ ДОП. ОБРАБОТКИ ###
def get_empty_pretrained_data(vector_size: int = 300) -> tuple[dict[str, int], list[str], np.ndarray, int]:
    if vector_size <= 0:
        raise ValueError('Значение vector_size должно быть положительным числом.\n')
    return {}, [], np.zeros((vector_size)), vector_size

def is_valid_folder_name(folder_name: str) -> bool:
    folder_name = folder_name.strip()

    # Проверка на пустую строку
    if not folder_name:
        return False

    # Недопустимые символы для Windows
    invalid_chars = '\\/:*?"<>|'
    if any(char in folder_name for char in invalid_chars):
        return False

    # Зарезервированные имена в Windows
    forbidden_names = {"CON", "PRN", "AUX", "NUL"} | {f"COM{i}" for i in range(1, 10)} | {f"LPT{i}" for i in range(1, 10)}
    if folder_name.upper() in forbidden_names:
        return False

    # Имя не должно начинаться или заканчиваться пробелом или точкой
    if folder_name[0] in {' ', '.'} or folder_name[-1] in {' ', '.'}:
        return False

    return True

def has_pt_files(dir_path) -> bool:
    return any(f.endswith('.pt') for f in os.listdir(dir_path) if os.path.isfile(os.path.join(dir_path, f)))

def convert_keys_to_str(d: dict) -> dict:
    return {f"{k[0]},{k[1]}": v.tolist() if isinstance(v, np.ndarray) else v for k, v in d.items()}

def convert_keys_to_tuple(d: dict) -> dict:
    return {tuple(map(int, k.split(','))): np.array(v) if isinstance(v, list) else v for k, v in d.items()}

def get_columns_from_file(file_path: str) -> list[str]:
    if not file_path.lower().endswith('.xlsb'):
        print_critical_error("Файл не имеет расширение .xlsb.", prefix='\n')
        print_critical_error(f"Путь к файлу: {file_path}", end='\n\n')
        exit(1)

    try:
        df = pd.read_excel(file_path, engine='pyxlsb', nrows=0)
        return list(df.columns)

    except Exception as e:
        print_critical_error("Ошибка при открытии или чтении файла.", prefix='\n')
        print_critical_error(f"Путь к файлу: {file_path}")
        print_critical_error(f"Причина: {e}", end='\n\n')
        exit(1)

def format_cluster_dict_multiline(d: dict[str, int]) -> str:
    return "; ".join(f"{k}: {v}%" for k, v in d.items())

def read_file_add_column_and_save(data_file_path: str, save_dir_path: str, request_clusters: list[dict[str,int]], target_column: str) -> Optional[str]:
    if not os.path.isfile(data_file_path):
        print_critical_error('Файл не найден.', prefix='\n')
        print_critical_error(f'Путь к файлу: {data_file_path}', end='\n\n')
        exit(1)
        
    if not os.path.isdir(save_dir_path):
        print_critical_error('Папка не найдена.', prefix='\n')
        print_critical_error(f'Путь: {save_dir_path}', end='\n\n')
        exit(1)

    if not data_file_path.endswith('.xlsb'):
        print_critical_error('Расширение файла должно быть .xlsb', prefix='\n')
        print_critical_error(f'Файл: {data_file_path}')
        exit(1)

    # Путь к сохраняемому файлу
    filename = os.path.basename(data_file_path)
    filename_no_ext = os.path.splitext(filename)[0]
    new_data_file_path = os.path.join(save_dir_path, f"clustered_{filename_no_ext}.xlsx")
    
    # Проверка существования сохраняемого файла
    if os.path.isfile(new_data_file_path):
        print_error('Сохраняемый файл уже существует!')
        print_error('Файл:', end=' '); print_key_info(f'{new_data_file_path}')
        print_warn('Удалите существующий файл или переименуйте файл с исходными данными во избежание перезаписи.')
        print_warn('Файл с исходными данными:', end=' '); print_key_info(f'{data_file_path}')
        print_sub_line()
        return None

    # Чтение файла
    df = pd.read_excel(data_file_path, engine='pyxlsb')
    
    # Стандартные классы
    standard_class_names = list(config.target_words_for_cluster_kohonen.values())
    # Подготовка колонок
    cluster_columns = {
        f'Кластер_хар_{i+1}': [] for i in range(len(standard_class_names))
    }
    cluster_columns['Кластер_Аномалия'] = []
    # Заполнение колонок
    cluster_idx = 0
    for value in df[target_column]:
        if pd.notna(value) and cluster_idx < len(request_clusters):
            # Название кластера для текущего обращения
            cluster = request_clusters[cluster_idx]
            # Разбор значений для стандартных классов из текущего кластера
            row_values = {class_name: f"{class_name}: {cluster[class_name]}%" for class_name in cluster if class_name in standard_class_names}
            # Добавление значений в соответствующие колонки для стандартных классов
            for i, class_name in enumerate(standard_class_names):
                col_name = f'Кластер_хар_{i+1}'
                cluster_columns[col_name].append(row_values.get(class_name, None))
            # Добавление значений в колонку для аномалий
            anomaly_parts = [f"{class_name}: {prob}%" for class_name, prob in cluster.items() if class_name not in standard_class_names]
            cluster_columns["Кластер_Аномалия"].append("; ".join(anomaly_parts) if anomaly_parts else None)
            
            cluster_idx += 1
        else:
            for col in cluster_columns:
                cluster_columns[col].append(None)
    # Добавление колонок в DataFrame
    for col_name, values in cluster_columns.items():
        df[col_name] = values

    # Используем openpyxl для настройки формата ячеек
    try:
        with pd.ExcelWriter(new_data_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

            # Доступ к открытому файлу
            workbook = writer.book
            sheet = workbook.active
        
            # Установка формата для столбца с датой и/или временем
            date_time_format = 'DD.MM.YYYY HH:MM:SS'
            date_format = 'DD.MM.YYYY'

            # Установка формата для столбца "Дата поступления"
            date_time_column = df.columns.get_loc('Дата поступления') + 1  # Нумерация столбцов начинается с 1
            for row in range(2, len(df) + 2):                              # Нумерация строк с 1; 1 строка - заголовок
                cell = sheet.cell(row=row, column=date_time_column)
                cell.number_format = date_time_format
                cell.alignment = Alignment(horizontal='left', wrap_text=True)

            # Установка формата для столбца "Дата закрытия"
            date_column = df.columns.get_loc('Дата закрытия') + 1  # Нумерация столбцов начинается с 1
            for row in range(2, len(df) + 2):                      # Нумерация строк с 1; 1 строка - заголовок
                cell = sheet.cell(row=row, column=date_column)
                cell.number_format = date_format
                cell.alignment = Alignment(horizontal='left', wrap_text=True)

            # Установка ширины столбцов
            column_widths = {
                'A': 12,
                'B': 20,
                'C': 15,
                'D': 25,
                'E': 10,
                'F': 13,
                'G': 14,
                'H': 10,
                'I': 10,
                'J': 10,
                'K': 12,
                'L': 10,
                'M': 12,
                'N': 12,
                'O': 20,
                'P': 30,
                'Q': 30,
                'R': 30,
                'S': 25,
                'T': 60
            }

            # Применение ширины к каждому столбцу
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Выравнивание по левому краю и жирный шрифт у заголовков
            for idx in range(1, len(df.columns) + 1):
                sheet.cell(row=1, column=idx).alignment = Alignment(horizontal='left')
                sheet.cell(row=1, column=idx).font = Font(name='Arial', size=10, bold=True)
            
            # Цвета и стили
            fill_standard = PatternFill(start_color='002060', end_color='002060', fill_type='solid')  # RGB(0,32,96)
            font_standard = Font(color='FFFFFF', bold=True, name='Arial', size=10)                    # Белый шрифт

            fill_other = PatternFill(start_color='92CDDC', end_color='92CDDC', fill_type='solid')     # RGB(146,205,220)
            font_other = Font(color='002060', bold=True, name='Arial', size=10)                       # RGB(0,32,96)

            # Форматирование заголовков
            for idx, col_name in enumerate(df.columns, start=1):
                # К добавленным колонкам применяется стиль "other"
                if col_name.startswith('Кластер_'):
                    sheet.cell(row=1, column=idx).fill = fill_other
                    sheet.cell(row=1, column=idx).font = font_other
                else: # Иначе стандартный
                    sheet.cell(row=1, column=idx).fill = fill_standard
                    sheet.cell(row=1, column=idx).font = font_standard

            # Установка Arial 10 для таблицы
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.font = Font(name='Arial', size=10)
    except PermissionError as e:
        print_critical_error('Невозможно сохранить файл.', prefix='\n')
        print_critical_error('Возможно, файл открыт в Excel или иная проблема.')
        print_critical_error(f'Путь к файлу: {new_data_file_path}')
        print_critical_error(f'Ошибка: {e}', end='\n\n')
        exit(1)
                
    return new_data_file_path

### ФУНКЦИИ ДЛЯ .JSON, .PKL ###
def save_to_pickle(data, file_path: str) -> None:
    """Сохраняет объект в .pkl файл по указанному пути."""
    try:
        with open(file_path, 'wb') as f:
            pickle.dump(data, f)
    except Exception as e:
        print_critical_error(f'Невозможно сохранить файл .pkl: {e}', end='\n\n', prefix='\n')
        exit(1)
    
def load_from_pickle(file_path: str):
    """Загружает объект из .pkl файла по указанному пути."""
    try:
        with open(file_path, 'rb') as f:
            return pickle.load(f)
    except Exception as e:
        print_critical_error(f'Невозможно загрузить файл .pkl: {e}', end='\n\n', prefix='\n')
        exit(1)
    
def save_to_json(data, file_path: str, ensure_ascii: bool = False, indent: int = 4) -> None:
    """Сохраняет объект в .json файл по указанному пути."""
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=ensure_ascii, indent=indent)
    except Exception as e:
        print_critical_error(f'Невозможно сохранить файл .json: {e}', end='\n\n', prefix='\n')
        exit(1)

def load_from_json(file_path: str):
    """Загружает объект из .json файла по указанному пути."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print_critical_error(f'Невозможно загрузить файл .json: {e}', end='\n\n', prefix='\n')
        exit(1)